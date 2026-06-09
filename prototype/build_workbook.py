import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
econ = json.load(open(os.path.join(ROOT, "data", "economy-model.json"), encoding="utf-8"))
cards = json.load(open(os.path.join(ROOT, "data", "cards.json"), encoding="utf-8"))

FONT = "Arial"
BLUE = Font(name=FONT, color="0000FF")        # inputs
BLACK = Font(name=FONT, color="000000")        # formulas
GREEN = Font(name=FONT, color="008000")        # cross-sheet links
HEAD = Font(name=FONT, bold=True, color="FFFFFF")
HFILL = PatternFill("solid", fgColor="305496")
TITLE = Font(name=FONT, bold=True, size=14)
NOTE = Font(name=FONT, italic=True, color="808080", size=9)
thin = Side(style="thin", color="D0D0D0")
BORD = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()
wb.remove(wb.active)

def sheet(name):
    return wb.create_sheet(name)

def header(ws, row, cols, widths=None):
    for i, c in enumerate(cols, 1):
        cell = ws.cell(row=row, column=i, value=c)
        cell.font = HEAD; cell.fill = HFILL; cell.alignment = Alignment(horizontal="center")
        cell.border = BORD
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

def title(ws, text):
    ws["A1"] = text; ws["A1"].font = TITLE

# ---------------- Constants ----------------
ws = sheet("Constants"); title(ws, "MULTIPLES — Constants (Layer-2 inputs)")
ws["A2"] = "Money: integer cents | Multiple: milli-units x1000 | Ownership: basis points x10000"; ws["A2"].font = NOTE
header(ws, 4, ["Constant", "Value", "Note"], [26, 16, 80])
r = 5
for k, v in econ["constants"].items():
    if k.startswith("_"):
        continue
    ws.cell(row=r, column=1, value=k).font = BLACK
    c = ws.cell(row=r, column=2, value=v); c.font = BLUE; c.fill = PatternFill("solid", fgColor="FFFDE7")
    ws.cell(row=r, column=3, value=econ["constants"].get("_note", "") if k == "_note" else "").font = NOTE
    r += 1
ws.cell(row=r+1, column=1, value="note").font = NOTE
ws.cell(row=r+1, column=3, value=econ["constants"].get("_note", "")).font = NOTE

# ---------------- Formulas ----------------
ws = sheet("Formulas"); title(ws, "Layer-1 FROZEN formulas (see resolver.dart)")
header(ws, 3, ["ID", "Definition"], [22, 110])
r = 4
for k, v in econ["formulas"].items():
    ws.cell(row=r, column=1, value=k).font = BLACK
    ws.cell(row=r, column=2, value=v).font = BLACK
    r += 1

# ---------------- Sectors ----------------
ws = sheet("Sectors"); title(ws, "Sectors (4) — base multiple seeds + volatility")
header(ws, 3, ["Sector", "baseMultiple (raw)", "as x", "Volatility", "Signature"], [14, 18, 10, 11, 80])
r = 4
for s in econ["sectors"]:
    ws.cell(row=r, column=1, value=s["name"]).font = BLACK
    ws.cell(row=r, column=2, value=s["baseMultiple"]).font = BLUE
    fx = ws.cell(row=r, column=3, value=f"=B{r}/1000"); fx.font = BLACK; fx.number_format = '0.0"x"'
    ws.cell(row=r, column=4, value=s["volatility"]).font = BLUE
    ws.cell(row=r, column=5, value=s["signature"]).font = BLACK
    r += 1

# ---------------- TierBars ----------------
ws = sheet("TierBars"); title(ws, "Tier net-worth bars (the ante track)")
header(ws, 3, ["Tier", "Display", "Bar (cents)", "Bar ($)", "Deadline (rounds)", "Floor clear %", "Who"], [7, 10, 16, 14, 16, 13, 40])
r = 4
for t in econ["tierBars"]:
    ws.cell(row=r, column=1, value=t["tier"]).font = BLACK
    ws.cell(row=r, column=2, value=t["barDisplay"]).font = BLACK
    ws.cell(row=r, column=3, value=t["bar"]).font = BLUE
    d = ws.cell(row=r, column=4, value=f"=C{r}/100"); d.font = BLACK; d.number_format = '$#,##0'
    ws.cell(row=r, column=5, value=t["deadlineRounds"]).font = BLUE
    p = ws.cell(row=r, column=6, value=t["floorClearPct"]); p.font = BLUE; p.number_format = '0.0%'
    ws.cell(row=r, column=7, value=t["who"]).font = BLACK
    r += 1

# ---------------- Curves ----------------
ws = sheet("Curves"); title(ws, "Curves — growth line, reinvest decay, interest, drift, leverage")
r = 3
ws.cell(row=r, column=1, value="Optimal growth (per tier)").font = Font(name=FONT, bold=True); r += 1
header(ws, r, ["Tier", "Growth needed", "Required/round", "Realistic early", "Realistic late"], [8, 14, 14, 14, 14]); r += 1
for g in econ["curves"]["optimalGrowth"]["perTierLine"]:
    ws.cell(row=r, column=1, value=g["tier"]).font = BLACK
    ws.cell(row=r, column=2, value=g.get("growthNeeded")).font = BLACK
    ws.cell(row=r, column=3, value=g.get("requiredPerRound")).font = BLACK
    ws.cell(row=r, column=4, value=g.get("realisticEarly")).font = BLACK
    ws.cell(row=r, column=5, value=g.get("realisticLate")).font = BLACK
    r += 1
r += 1
ws.cell(row=r, column=1, value="Reinvest-efficiency decay").font = Font(name=FONT, bold=True); r += 1
header(ws, r, ["Progress", "Efficiency"], [12, 12]); r += 1
for s in econ["curves"]["reinvestDecay"]["samples"]:
    ws.cell(row=r, column=1, value=s["progress"]).font = BLACK
    e = ws.cell(row=r, column=2, value=s["eff"]); e.font = BLACK; e.number_format = '0%'
    r += 1
r += 1
ws.cell(row=r, column=1, value="Interest band").font = Font(name=FONT, bold=True); r += 1
header(ws, r, ["State", "Min", "Max", "Rate mult", "Financing"], [10, 8, 8, 10, 60]); r += 1
for k in ("normal", "bubble", "crunch"):
    b = econ["curves"]["interestBand"][k]
    ws.cell(row=r, column=1, value=k).font = BLACK
    ws.cell(row=r, column=2, value=b["min"]).font = BLUE
    ws.cell(row=r, column=3, value=b["max"]).font = BLUE
    ws.cell(row=r, column=4, value=b["rateMul"]).font = BLUE
    ws.cell(row=r, column=5, value=b["financing"]).font = NOTE
    r += 1

# ---------------- Plays ----------------
ws = sheet("Plays"); title(ws, "PLAYS — the one-shot consumable layer (§8 Q2)")
header(ws, 3, ["Play", "Gate", "Deltas / effect", "Warning"], [18, 10, 70, 40])
r = 4
for k, v in econ["plays"].items():
    ws.cell(row=r, column=1, value=k).font = BLACK
    ws.cell(row=r, column=2, value=v.get("gate", "")).font = BLUE
    ws.cell(row=r, column=3, value=v.get("deltas", "")).font = BLACK
    ws.cell(row=r, column=4, value=v.get("warning", "")).font = NOTE
    r += 1

# ---------------- TuningKnobs ----------------
ws = sheet("TuningKnobs"); title(ws, "Tuning knobs (turn these first)")
header(ws, 3, ["Knob", "Value", "Note"], [24, 18, 80])
r = 4
for t in econ["tuningKnobs"]:
    ws.cell(row=r, column=1, value=t["name"]).font = BLACK
    val = t["value"]
    c = ws.cell(row=r, column=2, value=json.dumps(val) if isinstance(val, list) else val)
    c.font = BLUE; c.fill = PatternFill("solid", fgColor="FFFDE7")
    ws.cell(row=r, column=3, value=t["note"]).font = NOTE
    r += 1

# ---------------- Cards ----------------
ws = sheet("Cards"); title(ws, f"Content / Card Database ({len(cards)} cards; vertical slice flagged)")
cols = ["id", "name", "type", "sector", "rarity", "tierGate", "cost",
        "d.ebitda", "d.multiple", "d.netDebt", "d.own", "d.cash", "VS?", "lesson", "flavor"]
header(ws, 3, cols, [16, 24, 11, 11, 9, 8, 24, 9, 10, 9, 7, 8, 6, 46, 46])
r = 4
for c in cards:
    d = c.get("deltas", {})
    cost = c.get("cost")
    cost_s = json.dumps(cost) if isinstance(cost, (dict, list)) else cost
    row = [c.get("id"), c.get("name"), c.get("type"), c.get("sector"), c.get("rarity"),
           c.get("tierGate"), cost_s,
           d.get("ebitda"), d.get("multiple"), d.get("netDebt"), d.get("own"), d.get("cash"),
           "Y" if c.get("inVerticalSlice") else "", c.get("lesson"), c.get("flavor")]
    for i, val in enumerate(row, 1):
        cell = ws.cell(row=r, column=i, value=val); cell.font = BLACK
        if i == 13 and val == "Y":
            cell.fill = PatternFill("solid", fgColor="E2EFDA"); cell.alignment = Alignment(horizontal="center")
    r += 1
ws.freeze_panes = "A4"

# ---------------- SimCheck ----------------
ws = sheet("SimCheck"); title(ws, "Sim-check (floor strategy, N=5000 seeded runs)")
sc = econ["simCheck"]
r = 3
ws.cell(row=r, column=1, value="Verdict").font = Font(name=FONT, bold=True)
ws.cell(row=r, column=2, value=sc["verdict"]).font = BLACK; r += 2
header(ws, r, ["Metric", "Value"], [28, 14]); r += 1
rows = [("Floor overall win rate", sc["floor"]["overallWinRate"]),
        ("Floor bankruptcy rate", sc["floor"]["bankruptcyRate"]),
        ("T1 clear %", sc["floor"]["perTierClearPct"]["T1"]),
        ("T2 clear %", sc["floor"]["perTierClearPct"]["T2"]),
        ("T3 clear %", sc["floor"]["perTierClearPct"]["T3"]),
        ("T4 clear %", sc["floor"]["perTierClearPct"]["T4"]),
        ("Greedy win rate", sc["greedy"]["winRate"]),
        ("Greedy bankruptcy rate", sc["greedy"]["bankruptcyRate"])]
for n, v in rows:
    ws.cell(row=r, column=1, value=n).font = BLACK
    c = ws.cell(row=r, column=2, value=v); c.font = BLACK; c.number_format = '0.0%'
    r += 1
ws.cell(row=r+1, column=1, value="Scope").font = NOTE
ws.cell(row=r+1, column=2, value=sc["scope"]).font = NOTE

# ---------------- Model (live, sim-able) ----------------
ws = sheet("Model"); title(ws, "Live model — edit the blue inputs, watch Net Worth recompute")
ws["A2"] = "Real in-cell formulas mirror resolver.dart. Single-venture worked example."; ws["A2"].font = NOTE
header(ws, 4, ["Input (blue = editable)", "Raw value", "Display"], [30, 18, 16])
inp = [("EBITDA (cents)", "=Constants!B6", econ["constants"]["startEbitda"]),
       ("Multiple (milli-units)", "=Constants!B8", econ["constants"]["startMultiple"]),
       ("Net Debt (cents)", "=Constants!B10", econ["constants"]["startNetDebt"]),
       ("Ownership (bp)", "=Constants!B9", econ["constants"]["startOwnership"]),
       ("Cash (cents)", "=Constants!B5", econ["constants"]["startCash"])]
# Use hardcoded blue inputs (editable scenario cells), not cross-sheet, so user can scenario freely
vals = {"EBITDA": econ["constants"]["startEbitda"], "Multiple": econ["constants"]["startMultiple"],
        "NetDebt": econ["constants"]["startNetDebt"], "Own": econ["constants"]["startOwnership"],
        "Cash": econ["constants"]["startCash"]}
labels = [("EBITDA (cents)", vals["EBITDA"]), ("Multiple (milli x1000)", vals["Multiple"]),
          ("Net Debt (cents)", vals["NetDebt"]), ("Ownership (bp x10000)", vals["Own"]),
          ("Cash (cents)", vals["Cash"])]
r = 5
for name, v in labels:
    ws.cell(row=r, column=1, value=name).font = BLACK
    c = ws.cell(row=r, column=2, value=v); c.font = BLUE; c.fill = PatternFill("solid", fgColor="FFFDE7")
    r += 1
# rows: B5 EBITDA, B6 Multiple, B7 NetDebt, B8 Own, B9 Cash
r += 1
ws.cell(row=r, column=1, value="Derived (black = formula)").font = Font(name=FONT, bold=True); r += 1
def out(label, formula, disp=None, money=False):
    global r
    ws.cell(row=r, column=1, value=label).font = BLACK
    c = ws.cell(row=r, column=2, value=formula); c.font = BLACK
    if disp:
        dd = ws.cell(row=r, column=3, value=disp); dd.font = GREEN; dd.number_format = '$#,##0'
    if money:
        ws.cell(row=r, column=3, value=f"=B{r}/100").number_format = '$#,##0'
        ws.cell(row=r, column=3).font = GREEN
    r += 1
out("Enterprise Value = trunc(EBITDA*Mult/1000)", "=TRUNC(B5*B6/1000)", money=True)
ev_row = r - 1
out("Equity = EV - NetDebt", f"=B{ev_row}-B7", money=True)
eq_row = r - 1
out("Net Worth = trunc(Own*Equity/10000)+Cash", f"=TRUNC(B8*B{eq_row}/10000)+B9", money=True)
out("Interest @14% = trunc(0.14*NetDebt)", "=TRUNC(0.14*B7)", money=True)
out("Bankrupt? (Cash-Interest < 0)", f"=IF(B9-TRUNC(0.14*B7)<0,\"YES\",\"no\")")
r += 1
ws.cell(row=r, column=1, value="Seed Net Worth should read $56,000 with default inputs.").font = NOTE

out_path = os.path.join(ROOT, "docs", "MULTIPLES-economy-and-content.xlsx")
wb.save(out_path)
print("SAVED", out_path)
print("TABS", wb.sheetnames)
